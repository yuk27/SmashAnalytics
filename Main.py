import pyautogui
import pytesseract
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import time
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import os
import datetime

amnt_characters = 84
amnt_players = 7

path = Path().absolute()
wb = Workbook()
pytesseract.pytesseract.tesseract_cmd = 'C:\\Program Files\\Tesseract-OCR\\tesseract'
file = None
player_name = None

character_info = {

    "Player": {
        "position": [1310, 137, 300, 50],  # X, Y, W, H
        "path": "{0}{1}".format(path, "\\img\\character_user.png"),
        "value": None
    },

    "character_name":
        {
            "position": [500, 165, 500, 100],
            "path": "{0}{1}".format(path, "\\img\\character.png"),
            "value": None
        },

    "KOs":
        {
            "position": [1173, 300, 120, 50],
            "path": "{0}{1}".format(path, "\\img\\character_ko.png"),
            "value": None
        },

    "falls":
        {
            "position": [1168, 769, 120, 50],
            "path": "{0}{1}".format(path, "\\img\\character_falls.png"),
            "value": None
        },

    "battles":
        {
            "position": [1168, 882, 120, 50],
            "path": "{0}{1}".format(path, "\\img\\character_battles.png"),
            "value": None
        }
}


def full_screenshot():
    fpath = "{0}{1}".format(path, "\\img\\screen.png")
    s = pyautogui.screenshot()
    s.save(fpath)


def take_screenshot(position, img_path):
    s = pyautogui.screenshot(region=(position[0], position[1], position[2], position[3]))
    s.save(img_path)


def img_to_text(position, img_path):
    take_screenshot(position, img_path)
    text = (pytesseract.image_to_string(img_path).strip())
    return text


def add_row(ws, r, values, offset_column=0, offset_row=0):
    for j in range(0, len(values)):
        if not str(values[j]).isnumeric():
            ws.cell(r+offset_row + 1, j + offset_column + 1, values[j])
        else:
            ws.cell(r + offset_row + 1, j + offset_column + 1, "=VALUE({0})".format(values[j]))

def press_key(key):
    pyautogui.keyDown(key)
    time.sleep(.001)
    pyautogui.keyUp(key)


gauth = GoogleAuth()

# Creates local webserver and auto
# handles authentication.
gauth.LocalWebserverAuth()
drive = GoogleDrive(gauth)

time.sleep(2)  # Sleep for 3 seconds
print("Start")


row = ["Player", "Character", "Kos", "Falls", "Battles", "Success Rate", "Fail rate",  "Won Battles"]

ws = wb.create_sheet()
ws.title = 'Export'  # player_name
add_row(ws, 0, row)

for y in range(0, amnt_players):
    press_key('down')
    player = img_to_text(character_info["Player"]["position"], character_info["Player"]["path"])

    for x in range(0, amnt_characters):

        character = img_to_text(character_info["character_name"]["position"], character_info["character_name"]["path"])
        kos = img_to_text(character_info["KOs"]["position"], character_info["KOs"]["path"])

        if kos == '' or kos is None:
            row = [player, character, 0, 0, 0, 0, 0, 0]

        else:
            press_key('down')
            press_key('down')
            press_key('down')
            press_key('down')
            press_key('down')

            falls = img_to_text(character_info["falls"]["position"], character_info["falls"]["path"])
            battles = img_to_text(character_info["battles"]["position"], character_info["battles"]["path"])

            press_key('up')
            press_key('up')
            press_key('up')
            press_key('up')
            press_key('up')

            kos = int(kos)
            if falls != "":
                falls = int(falls)

                if kos + falls != 0:
                    success_rate = kos / (kos + falls)
                else:
                    success_rate = 0

                if battles != '' and battles is not None:
                    won_battles = int(battles) * success_rate
                else:
                    won_battles = 0

            row = [
                player,
                character,
                kos,
                falls,
                battles,
                success_rate,
                1 - success_rate,
                won_battles
            ]

        add_row(ws, x + (y * amnt_characters) + 1, row)

        # Move to next character
        press_key('e')

        print("Player{2} - character {0}/{3}: {1} - {4}".format(x+1, row, y, amnt_characters, (x + (y * amnt_characters))))

    # Next player
    # player_name = None
    press_key('x')
    press_key('x')
    press_key('down')
    press_key('c')
    press_key('c')
    time.sleep(.2)


tab = Table(displayName="Table1", ref="A1:H{0}".format(amnt_characters * amnt_players+1))


# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style

ws.add_table(tab)

std = wb['Sheet']
wb.remove(std)

wb.save(filename='{0}\\{1}'.format(path, "stats.xlsx"))


# iterating thought all the files/folder
# of the desired directory
for x in os.listdir(path):
    if x[-4:] == 'xlsx':
        f = drive.CreateFile({'title': "{0}_{1}".format(datetime.datetime.now(), x)})
        f.SetContentFile(os.path.join(path, x))
        f.Upload()
